import camelot
import pandas as pd
from openpyxl import Workbook, load_workbook
import re
def convert_pdf_and_filter(pdf_path, output_file):
 
    # Convert PDF to Excel
    tables = camelot.read_pdf(pdf_path, flavor='stream', pages='all')
    dfs = [table.df for table in tables]
    final_df = pd.concat(dfs, ignore_index=True)
 
    # Save to Excel without filtering
    final_df.to_excel(output_file, index=False, header=False)
 
    # Read the Excel file
    wb = load_workbook(output_file)
    ws = wb.active
 
    # Remove backslash from all cells
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                cell.value = str(cell.value).replace('\\', '')
                cell.value = str(cell.value).replace('数量', '')
                cell.value = str(cell.value).replace('単価(税抜き)', '')
 
    # Remove columns and rows
    ws.delete_cols(1, 2)
    ws.delete_cols(4, 3)
    ws.delete_rows(1, 5)
 
    # Apply regex replacement to column 4
    for row in ws.iter_rows(min_row=2, max_col=4, max_row=ws.max_row):
        for cell in row:
            if cell.column == 4 and cell.value is not None:
                cell.value = re.sub(r'\D', '', str(cell.value))
 
    # Add SUM formula to cell I5
    ws['I5'] = '=SUMPRODUCT(B2:B200, C2:C200)'
 
    # Save the final Excel file
    wb.save(output_file)
if __name__ == "__main__":
    pdf_file_path = 'Downloads/quote.pdf'
    output_file_path = 'Downloads/quote.xlsx'
    convert_pdf_and_filter(pdf_file_path, output_file_path)
    print(f"Filtered data saved to {output_file_path}")
